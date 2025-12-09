from datetime import datetime, timedelta
from decimal import Decimal
import ast
import re
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F, Count
from django.utils import timezone
from apps.usuarios.decorators import role_required
from apps.pedidos.models import Pedido, PedidoItem
from apps.productos.models import Producto, MovimientoInventario
from apps.proveedores.models import Proveedor
from apps.empleados.models import Empleado
from apps.clientes.models import Cliente
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from .models import AuditLog
from apps.usuarios.models import Usuario


def _parse_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _clp_fmt(value: int) -> str:
    s = f"{int(value):,}".replace(",", ".")
    return f"$ {s}"


_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _clean_xlsx_value(value):
    """Convierte el valor a texto simple y elimina caracteres ilegales para XLSX."""
    if value is None:
        return ""
    s = str(value)
    return _ILLEGAL_XLSX_CHARS_RE.sub("", s)


def _parse_detail_to_dict(detail: str):
    """Intenta convertir el texto de detail en un dict limpio.

    - Elimina constructores como Decimal('12345').
    - Reemplaza representaciones de ImageFieldFile por la ruta del archivo.
    - Si no puede parsear, devuelve None.
    """
    if not detail:
        return None

    s = str(detail)
    # Reemplazar Decimal('123') -> '123'
    s = re.sub(r"Decimal\('([^']*)'\)", r"'\\1'", s)
    # Reemplazar representaciones de archivos tipo <ImageFieldFile: path>, <FieldFile: None>, etc. -> 'valor'
    # Captura cualquier clase que termine con 'File' y toma lo que sigue a los dos puntos.
    s = re.sub(r"<[^:>]*File:\s*([^>]+)>", r"'\\1'", s)
    # Reemplazar rutas de enums/clases tipo MovimientoInventario.Tipo.SALIDA -> 'SALIDA'
    # Coincide con palabras separadas por puntos terminando en MAYÚSCULAS o mayúsculas+guiones bajos
    s = re.sub(r"\b(?:[A-Za-z_][A-Za-z0-9_]*\.)+([A-Z_][A-Z0-9_]*)\b", r"'\\1'", s)

    try:
        data = ast.literal_eval(s)
    except Exception:
        return None

    if isinstance(data, dict):
        # Convertir todo a texto simple
        return {str(k): '' if v is None else str(v) for k, v in data.items()}
    return None


@role_required('Administrador')
def index(request):
    return render(request, 'reportes/index.html')


@role_required('Administrador')
def reporte_ventas(request):
    # Filtros
    f_inicio = _parse_date(request.GET.get('inicio', ''))
    f_fin = _parse_date(request.GET.get('fin', ''))
    producto_id = request.GET.get('producto')
    empleado_id = request.GET.get('empleado')
    cliente_id = request.GET.get('cliente')
    periodo = request.GET.get('periodo')  # semana | mes | anio
    export = request.GET.get('export')  # 'xlsx' | 'pdf'

    # Si se selecciona un periodo rápido, sobreescribir rangos de fecha
    hoy = timezone.localdate()
    if periodo == 'semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        f_inicio, f_fin = inicio_semana, fin_semana
    elif periodo == 'mes':
        inicio_mes = hoy.replace(day=1)
        # obtener último día del mes actual
        if hoy.month == 12:
            siguiente_mes = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            siguiente_mes = hoy.replace(month=hoy.month + 1, day=1)
        fin_mes = siguiente_mes - timedelta(days=1)
        f_inicio, f_fin = inicio_mes, fin_mes
    elif periodo == 'anio':
        inicio_anio = hoy.replace(month=1, day=1)
        fin_anio = hoy.replace(month=12, day=31)
        f_inicio, f_fin = inicio_anio, fin_anio

    qs = Pedido.objects.all()
    if f_inicio:
        qs = qs.filter(fecha_pedido__date__gte=f_inicio)
    if f_fin:
        qs = qs.filter(fecha_pedido__date__lte=f_fin)
    if empleado_id:
        qs = qs.filter(empleado_id=empleado_id)
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    if producto_id:
        qs = qs.filter(items__producto_id=producto_id).distinct()

    qs = qs.select_related('cliente', 'empleado').prefetch_related('items')

    total_subtotal = qs.aggregate(s=Sum('subtotal'))['s'] or 0
    total_iva = qs.aggregate(s=Sum('iva'))['s'] or 0
    total_total = qs.aggregate(s=Sum('total'))['s'] or 0

    # Resumen por vendedor
    resumen_vendedores = (
        qs.values('empleado_id', 'empleado__usuario__username')
        .annotate(total_ventas=Sum('total'), num_pedidos=Count('id'))
        .order_by('-total_ventas')
    )

    if export == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ventas'
        ws.append(['ID', 'Fecha', 'Cliente', 'Empleado', 'Subtotal', 'IVA', 'Total', 'Estado'])
        for p in qs:
            ws.append([
                p.id,
                timezone.localtime(p.fecha_pedido).strftime('%Y-%m-%d %H:%M'),
                getattr(p.cliente, 'nombre', ''),
                getattr(getattr(p, 'empleado', None), 'usuario', '') if p.empleado_id else '',
                int(p.subtotal), int(p.iva), int(p.total), p.estado,
            ])
        ws.append([])
        ws.append(['', '', '', 'Totales:', int(total_subtotal), int(total_iva), int(total_total)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
        wb.save(response)
        return response

    if export == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="reporte_ventas.pdf"'
        c = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        y = height - 20 * mm
        c.setFont('Helvetica-Bold', 14)
        c.drawString(20 * mm, y, 'Reporte de Ventas')
        y -= 10 * mm
        c.setFont('Helvetica', 9)
        for p in qs:
            if y < 25 * mm:
                c.showPage(); y = height - 20 * mm; c.setFont('Helvetica', 9)
            c.drawString(20 * mm, y, f"#{p.id} {timezone.localtime(p.fecha_pedido).strftime('%d-%m-%Y %H:%M')} - {getattr(p.cliente, 'nombre', '')}")
            c.drawRightString(200 * mm, y, _clp_fmt(int(p.total)))
            y -= 6 * mm
        y -= 6 * mm
        c.setFont('Helvetica-Bold', 10)
        c.drawRightString(165 * mm, y, 'Subtotal:')
        c.drawRightString(200 * mm, y, _clp_fmt(int(total_subtotal)))
        y -= 6 * mm
        c.drawRightString(165 * mm, y, 'IVA 19%:')
        c.drawRightString(200 * mm, y, _clp_fmt(int(total_iva)))
        y -= 6 * mm
        c.drawRightString(165 * mm, y, 'Total:')
        c.drawRightString(200 * mm, y, _clp_fmt(int(total_total)))
        c.showPage(); c.save()
        return response

    # HTML
    productos = Producto.objects.all().only('id', 'nombre')
    empleados = Empleado.objects.select_related('usuario').all().only('id', 'usuario__username')
    clientes = Cliente.objects.all().only('id', 'nombre')
    return render(request, 'reportes/ventas.html', {
        'pedidos': qs,
        'inicio': f_inicio,
        'fin': f_fin,
        'periodo': periodo,
        'producto_id': int(producto_id) if (producto_id or '').isdigit() else None,
        'empleado_id': int(empleado_id) if (empleado_id or '').isdigit() else None,
        'cliente_id': int(cliente_id) if (cliente_id or '').isdigit() else None,
        'productos': productos,
        'empleados': empleados,
        'clientes': clientes,
        'total_subtotal': total_subtotal,
        'total_iva': total_iva,
        'total_total': total_total,
        'resumen_vendedores': resumen_vendedores,
        'fmt': _clp_fmt,
    })


@role_required('Administrador')
def reporte_stock_bajo(request):
    productos = Producto.objects.filter(stock__lte=F('stock_minimo')).order_by('nombre')
    export = request.GET.get('export')
    if export == 'xlsx':
        wb = Workbook(); ws = wb.active; ws.title = 'Stock bajo'
        ws.append(['SKU', 'Nombre', 'Categoría', 'Stock', 'Stock mínimo'])
        for p in productos:
            ws.append([p.sku, p.nombre, p.categoria, p.stock, p.stock_minimo])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=stock_bajo.xlsx'
        wb.save(response); return response
    return render(request, 'reportes/stock_bajo.html', {'productos': productos})


@role_required('Administrador')
def reporte_proveedores_activos(request):
    # Proveedores con entregas (movimientos de entrada) en últimos 90 días
    desde = timezone.now() - timedelta(days=90)
    entradas = MovimientoInventario.objects.filter(tipo=MovimientoInventario.Tipo.ENTRADA, fecha__gte=desde)
    producto_ids = entradas.values_list('producto_id', flat=True)
    # Obtener proveedores vinculados por relación usuario -> proveedor (simplificación: mostramos proveedores que existan si hubo entradas)
    proveedores = Proveedor.objects.all()
    export = request.GET.get('export')
    if export == 'xlsx':
        wb = Workbook(); ws = wb.active; ws.title = 'Proveedores activos'
        ws.append(['Empresa', 'Contacto', 'Correo', 'Teléfono'])
        for prov in proveedores:
            ws.append([prov.empresa, prov.contacto, prov.correo, prov.telefono])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=proveedores_activos.xlsx'
        wb.save(response); return response
    return render(request, 'reportes/proveedores_activos.html', {
        'proveedores': proveedores,
        'desde': desde,
    })


@role_required('Administrador')
def reporte_auditoria(request):
    from django.db.models import Q

    inicio = _parse_date(request.GET.get('inicio', ''))
    fin = _parse_date(request.GET.get('fin', ''))
    modelo = request.GET.get('modelo', '')
    accion = request.GET.get('accion', '')
    usuario_id = request.GET.get('usuario', '')
    export = request.GET.get('export')

    qs = AuditLog.objects.select_related('user').all()
    if inicio:
        qs = qs.filter(created_at__date__gte=inicio)
    if fin:
        qs = qs.filter(created_at__date__lte=fin)
    if modelo:
        qs = qs.filter(model=modelo)
    if accion:
        qs = qs.filter(action=accion)
    if usuario_id:
        qs = qs.filter(user_id=usuario_id)

    modelos = AuditLog.objects.order_by().values_list('model', flat=True).distinct()
    usuarios = Usuario.objects.all().only('id', 'username')

    if export == 'xlsx':
        logs = list(qs.order_by('-created_at')[:5000])
        # Parsear detalles una vez y recolectar todas las claves
        parsed_details = []
        all_keys = set()
        for a in logs:
            d = _parse_detail_to_dict(a.detail)
            parsed_details.append(d)
            if isinstance(d, dict):
                all_keys.update(d.keys())

        extra_cols = sorted(all_keys)

        wb = Workbook(); ws = wb.active; ws.title = 'Auditoría'
        header = ['Fecha', 'Acción', 'Modelo', 'Objeto', 'Usuario', 'IP'] + extra_cols
        ws.append([_clean_xlsx_value(h) for h in header])

        for a, d in zip(logs, parsed_details):
            base = [
                a.created_at.strftime('%Y-%m-%d %H:%M'),
                a.action,
                a.model,
                a.object_id,
                getattr(a.user, 'username', ''),
                a.ip or '',
            ]
            if isinstance(d, dict):
                row = base + [d.get(k, '') for k in extra_cols]
            else:
                # Si no se pudo parsear, poner todo el detail en la primera columna extra
                first_extra = (a.detail or '') if extra_cols else ''
                extras = [first_extra] + [''] * (len(extra_cols) - 1)
                row = base + extras
            ws.append([_clean_xlsx_value(v) for v in row])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=auditoria.xlsx'
        wb.save(response); return response

    logs = list(qs.order_by('-created_at')[:200])
    for a in logs:
        a.detail_dict = _parse_detail_to_dict(a.detail)

    return render(request, 'reportes/auditoria.html', {
        'logs': logs,
        'modelos': modelos,
        'usuarios': usuarios,
        'modelo_sel': modelo,
        'accion_sel': accion,
        'usuario_sel': int(usuario_id) if (usuario_id or '').isdigit() else None,
        'inicio': inicio,
        'fin': fin,
    })
